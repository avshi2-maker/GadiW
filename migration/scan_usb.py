"""
GadiW Migration — Step 1: USB Inventory Scanner
================================================
Scans a folder (typically a USB drive) and produces an Excel inventory file.
The human (Avshi + Gadi) then fills in client/tag/direction columns in Excel.
The filled Excel is fed to migrate_to_gadiw.py for the actual upload.

Usage:
    python scan_usb.py <source_folder>

Examples:
    python scan_usb.py D:\
    python scan_usb.py C:\dev\gadiV\migration\test_files
"""

import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.table import Table

console = Console()

# ============================================================================
# CONFIGURATION
# ============================================================================

# File types we care about (Gadi's legal documents)
ALLOWED_EXTENSIONS = {
    '.pdf', '.docx', '.doc',
    '.xlsx', '.xls',
    '.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.tiff', '.tif',
    '.txt', '.rtf',
    '.eml', '.msg',  # email files
}

# File names/patterns to skip (system junk)
SKIP_FILE_NAMES = {
    '.ds_store', 'thumbs.db', 'desktop.ini', '.git', '.gitignore',
}

# File patterns to skip (start with these strings)
SKIP_PREFIXES = {'.', '~$', '$RECYCLE'}

# Max file size (50MB — same as Supabase Storage limit)
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024

# Direction enum values (matching documents.direction column)
DIRECTION_VALUES = ['פנימי', 'התקבל', 'נשלח']


# ============================================================================
# CORE SCANNING
# ============================================================================

def should_skip_file(file_path: Path) -> tuple[bool, str]:
    """Returns (skip, reason). reason is a human-readable string if skipping."""
    name = file_path.name.lower()
    
    # Skip by exact name match
    if name in SKIP_FILE_NAMES:
        return True, "system file"
    
    # Skip by prefix
    for prefix in SKIP_PREFIXES:
        if name.startswith(prefix):
            return True, f"hidden/temp ({prefix})"
    
    # Skip by extension
    ext = file_path.suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return True, f"unsupported type ({ext or 'no extension'})"
    
    # Skip if too large
    try:
        size = file_path.stat().st_size
        if size > MAX_FILE_SIZE_BYTES:
            return True, f"too large ({size // (1024*1024)} MB)"
        if size == 0:
            return True, "empty file"
    except OSError as e:
        return True, f"unreadable: {e}"
    
    return False, ""


def get_folder_hint(file_path: Path, source_root: Path) -> str:
    """Returns the folder path relative to source root (excluding the filename)."""
    try:
        rel = file_path.relative_to(source_root)
        parent = rel.parent
        return str(parent) if str(parent) != '.' else '(root)'
    except ValueError:
        return '(unknown)'


def scan_folder(source_root: Path):
    """
    Walk the source folder and return:
      - list of dicts with file metadata (for files we'll keep)
      - list of skipped files with reasons (for the report)
    """
    kept = []
    skipped = []
    
    # First pass: count total candidates for progress bar
    console.print(f"[cyan]Counting files in {source_root}...[/cyan]")
    all_files = []
    for root, dirs, files in os.walk(source_root):
        # Skip hidden/system directories
        dirs[:] = [d for d in dirs if not d.startswith('.') and d not in {'$RECYCLE.BIN', 'System Volume Information'}]
        for filename in files:
            all_files.append(Path(root) / filename)
    
    console.print(f"[green]Found {len(all_files)} candidate files. Processing...[/green]\n")
    
    # Second pass: classify each file
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Scanning files...", total=len(all_files))
        
        for file_path in all_files:
            progress.advance(task)
            
            skip, reason = should_skip_file(file_path)
            if skip:
                skipped.append({
                    'path': str(file_path),
                    'name': file_path.name,
                    'reason': reason,
                })
                continue
            
            try:
                stat = file_path.stat()
                kept.append({
                    'file_path': str(file_path),
                    'file_name': file_path.name,
                    'file_size_kb': round(stat.st_size / 1024, 1),
                    'folder_hint': get_folder_hint(file_path, source_root),
                    'file_modified_date': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d'),
                    'extension': file_path.suffix.lower(),
                })
            except OSError as e:
                skipped.append({
                    'path': str(file_path),
                    'name': file_path.name,
                    'reason': f"stat failed: {e}",
                })
    
    return kept, skipped


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def generate_excel(kept_files, skipped_files, source_root: Path, output_path: Path):
    """Create the inventory Excel with 3 sheets."""
    wb = Workbook()
    
    # ----- SHEET 1: Files (the main inventory) -----
    ws_files = wb.active
    ws_files.title = "Files"
    ws_files.sheet_view.rightToLeft = True  # RTL because Hebrew metadata
    
    headers = [
        'row_id',           # A — auto-numbered
        'file_path',        # B — full path on disk (read-only reference)
        'file_name',        # C — just the filename
        'file_size_kb',     # D — auto-detected
        'folder_hint',      # E — folder relative to USB root (helps fill client)
        'file_modified_date', # F — from filesystem
        'client_name',      # G — HUMAN FILLS (Hebrew client name as in DB)
        'doc_tag',          # H — HUMAN FILLS (חוזה/מכתב/תעודה etc.)
        'direction',        # I — HUMAN FILLS (פנימי/התקבל/נשלח dropdown)
        'doc_date',         # J — HUMAN FILLS or leave blank → uses file_modified_date
        'description',      # K — HUMAN FILLS (optional)
        'status',           # L — populated by migrate_to_gadiw.py later
    ]
    
    # Write header row with styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill_auto = PatternFill('solid', fgColor='1F3A5F')      # navy — auto-filled cols
    header_fill_human = PatternFill('solid', fgColor='2E7D32')     # green — HUMAN FILLS
    header_fill_status = PatternFill('solid', fgColor='888888')    # grey — script writes
    
    auto_cols = {'row_id', 'file_path', 'file_name', 'file_size_kb', 'folder_hint', 'file_modified_date'}
    human_cols = {'client_name', 'doc_tag', 'direction', 'doc_date', 'description'}
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_files.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if header in auto_cols:
            cell.fill = header_fill_auto
        elif header in human_cols:
            cell.fill = header_fill_human
        else:
            cell.fill = header_fill_status
    
    # Freeze header row
    ws_files.freeze_panes = 'A2'
    
    # Write data rows
    for idx, file_info in enumerate(kept_files, 2):
        ws_files.cell(row=idx, column=1, value=idx - 1)  # row_id
        ws_files.cell(row=idx, column=2, value=file_info['file_path'])
        ws_files.cell(row=idx, column=3, value=file_info['file_name'])
        ws_files.cell(row=idx, column=4, value=file_info['file_size_kb'])
        ws_files.cell(row=idx, column=5, value=file_info['folder_hint'])
        ws_files.cell(row=idx, column=6, value=file_info['file_modified_date'])
        # G-K stay blank for human to fill
        # L (status) stays blank — migrate script writes here
    
    # Add direction dropdown (column I) — apply to all data rows
    if kept_files:
        dv = DataValidation(
            type='list',
            formula1=f'"{",".join(DIRECTION_VALUES)}"',
            allow_blank=True,
        )
        dv.error = 'יש לבחור: פנימי, התקבל, או נשלח'
        dv.errorTitle = 'ערך לא חוקי'
        ws_files.add_data_validation(dv)
        last_row = len(kept_files) + 1
        dv.add(f'I2:I{last_row}')
    
    # Column widths
    widths = {
        'A': 8,    # row_id
        'B': 50,   # file_path
        'C': 30,   # file_name
        'D': 12,   # size_kb
        'E': 20,   # folder_hint
        'F': 14,   # modified_date
        'G': 24,   # client_name (human fill)
        'H': 14,   # doc_tag (human fill)
        'I': 12,   # direction (human fill)
        'J': 14,   # doc_date (human fill)
        'K': 30,   # description (human fill)
        'L': 14,   # status
    }
    for letter, width in widths.items():
        ws_files.column_dimensions[letter].width = width
    
    # ----- SHEET 2: Instructions -----
    ws_inst = wb.create_sheet("Instructions")
    ws_inst.sheet_view.rightToLeft = True
    
    instructions = [
        ('GadiW Migration — Excel Filling Instructions', True),
        ('הוראות מילוי קובץ Excel — מיגרציית GadiW', True),
        ('', False),
        ('English:', True),
        ('1. Fill ONLY the GREEN columns (G–K). Navy columns are auto-detected, do not edit.', False),
        ('2. client_name (G) — type the EXACT Hebrew name as it appears in the GadiW system.', False),
        ('   Tip: Sort by folder_hint (column E), then bulk-fill rows from the same folder.', False),
        ('3. doc_tag (H) — short Hebrew tag like חוזה / מכתב / תעודה / קבלה.', False),
        ('4. direction (I) — pick from dropdown: פנימי, התקבל, or נשלח.', False),
        ('5. doc_date (J) — date of the document (YYYY-MM-DD). If blank, file modified date is used.', False),
        ('6. description (K) — optional free-text notes.', False),
        ('7. status (L) — leave blank. The migration script writes here.', False),
        ('', False),
        ('עברית:', True),
        ('1. מלא רק את העמודות הירוקות (G–K). העמודות בכחול נקבעות אוטומטית.', False),
        ('2. client_name — שם הלקוח כפי שמופיע במערכת GadiW.', False),
        ('   טיפ: מיין לפי folder_hint, ומלא בקבוצות לפי תיקייה.', False),
        ('3. doc_tag — תיוג קצר: חוזה / מכתב / תעודה / קבלה.', False),
        ('4. direction — בחר מהרשימה: פנימי / התקבל / נשלח.', False),
        ('5. doc_date — תאריך המסמך. ריק = משתמש בתאריך מהקובץ.', False),
        ('6. description — תיאור חופשי (אופציונלי).', False),
        ('7. status — השאר ריק. סקריפט המיגרציה ימלא.', False),
        ('', False),
        ('Pro tips:', True),
        ('- Excel "Sort" by folder_hint → bulk-fill same-client rows fast', False),
        ('- Excel "Filter" by extension → handle all PDFs together, then all DOCX, etc.', False),
        ('- Excel "Find & Replace" inside one column to fix typos across many rows', False),
        ('- Save often (Ctrl+S). Excel does NOT auto-save .xlsx files.', False),
    ]
    
    for idx, (text, is_bold) in enumerate(instructions, 1):
        cell = ws_inst.cell(row=idx, column=1, value=text)
        if is_bold:
            cell.font = Font(bold=True, size=12)
    
    ws_inst.column_dimensions['A'].width = 100
    
    # ----- SHEET 3: Valid Values (reference) -----
    ws_ref = wb.create_sheet("Valid Values")
    ws_ref.sheet_view.rightToLeft = True
    ws_ref.cell(row=1, column=1, value='direction values:').font = Font(bold=True)
    for idx, val in enumerate(DIRECTION_VALUES, 2):
        ws_ref.cell(row=idx, column=1, value=val)
    
    ws_ref.cell(row=1, column=3, value='Common doc_tag suggestions:').font = Font(bold=True)
    common_tags = ['חוזה', 'מכתב', 'תעודה', 'קבלה', 'חשבונית', 'דוח', 'תמונה', 'אישור']
    for idx, val in enumerate(common_tags, 2):
        ws_ref.cell(row=idx, column=3, value=val)
    
    ws_ref.column_dimensions['A'].width = 20
    ws_ref.column_dimensions['C'].width = 30
    
    # ----- Save -----
    wb.save(output_path)


# ============================================================================
# REPORT
# ============================================================================

def print_summary(kept_files, skipped_files, source_root: Path, output_path: Path):
    """Pretty terminal summary."""
    console.print()
    
    # Summary table
    table = Table(title="Scan Summary", show_header=True, header_style="bold cyan")
    table.add_column("Metric", style="cyan")
    table.add_column("Count", justify="right", style="green")
    
    table.add_row("Source folder", str(source_root))
    table.add_row("Files kept (in Excel)", str(len(kept_files)))
    table.add_row("Files skipped", str(len(skipped_files)))
    table.add_row("Total examined", str(len(kept_files) + len(skipped_files)))
    table.add_row("Output Excel", str(output_path))
    
    console.print(table)
    console.print()
    
    # Skip reasons breakdown
    if skipped_files:
        skip_reasons = {}
        for sk in skipped_files:
            reason = sk['reason']
            skip_reasons[reason] = skip_reasons.get(reason, 0) + 1
        
        skip_table = Table(title="Skip Reasons", show_header=True, header_style="bold yellow")
        skip_table.add_column("Reason", style="yellow")
        skip_table.add_column("Count", justify="right")
        for reason, count in sorted(skip_reasons.items(), key=lambda x: -x[1]):
            skip_table.add_row(reason, str(count))
        console.print(skip_table)
        console.print()
    
    # File type breakdown
    if kept_files:
        ext_counts = {}
        for f in kept_files:
            ext_counts[f['extension']] = ext_counts.get(f['extension'], 0) + 1
        
        ext_table = Table(title="File Types Kept", show_header=True, header_style="bold green")
        ext_table.add_column("Extension", style="green")
        ext_table.add_column("Count", justify="right")
        for ext, count in sorted(ext_counts.items(), key=lambda x: -x[1]):
            ext_table.add_row(ext, str(count))
        console.print(ext_table)
        console.print()
    
    # Total size
    total_kb = sum(f['file_size_kb'] for f in kept_files)
    if total_kb > 1024 * 1024:
        size_str = f"{total_kb / (1024*1024):.2f} GB"
    elif total_kb > 1024:
        size_str = f"{total_kb / 1024:.2f} MB"
    else:
        size_str = f"{total_kb:.0f} KB"
    console.print(f"[bold]Total size of kept files:[/bold] {size_str}")
    console.print()


# ============================================================================
# MAIN
# ============================================================================

def main():
    if len(sys.argv) < 2:
        console.print("[red]ERROR: Source folder required.[/red]")
        console.print("[yellow]Usage: python scan_usb.py <source_folder>[/yellow]")
        console.print("[yellow]Example: python scan_usb.py D:\\[/yellow]")
        sys.exit(1)
    
    source_str = sys.argv[1]
    source_root = Path(source_str).resolve()
    
    # Validate source
    if not source_root.exists():
        console.print(f"[red]ERROR: Source folder does not exist: {source_root}[/red]")
        sys.exit(1)
    if not source_root.is_dir():
        console.print(f"[red]ERROR: Source is not a folder: {source_root}[/red]")
        sys.exit(1)
    
    # Warn if source looks like a USB drive (per Decision 5 — warn but allow)
    if len(str(source_root)) <= 3 and str(source_root)[1:3] == ':\\':
        drive_letter = str(source_root)[0].upper()
        if drive_letter not in ('C',):  # C: is usually the system drive
            console.print(f"[yellow]⚠ WARNING: Source looks like a removable drive ({source_root}).[/yellow]")
            console.print(f"[yellow]  Recommended: copy files to local hard drive first, then run from there.[/yellow]")
            console.print(f"[yellow]  Reason: USB drives can disconnect mid-migration causing partial uploads.[/yellow]")
            response = input("\n  Continue anyway? [y/N]: ").strip().lower()
            if response != 'y':
                console.print("[red]Aborted by user.[/red]")
                sys.exit(0)
    
    console.print()
    console.rule("[bold cyan]GadiW USB Inventory Scanner[/bold cyan]")
    console.print(f"Source: [bold]{source_root}[/bold]")
    console.print()
    
    # Scan
    kept_files, skipped_files = scan_folder(source_root)
    
    if not kept_files:
        console.print("[red]No files to migrate found.[/red]")
        sys.exit(0)
    
    # Generate Excel
    # Israel time (UTC+2 winter, UTC+3 summer DST)
from datetime import timezone, timedelta
israel_offset = timedelta(hours=3) if datetime.now().month in [3,4,5,6,7,8,9,10] else timedelta(hours=2)
israel_tz = timezone(israel_offset)
timestamp = datetime.now(israel_tz).strftime('%d%m%Y_%H%M%S')
    output_path = Path.cwd() / f'gadi_inventory_{timestamp}.xlsx'
    
    console.print(f"[cyan]Generating Excel: {output_path.name}[/cyan]")
    generate_excel(kept_files, skipped_files, source_root, output_path)
    console.print(f"[green]✓ Excel saved.[/green]")
    
    # Summary
    print_summary(kept_files, skipped_files, source_root, output_path)
    
    console.print("[bold green]Next step:[/bold green] open the Excel and fill columns G–K (green headers).")
    console.print("[bold green]Then run:[/bold green] python migrate_to_gadiw.py --sheet " + output_path.name)


if __name__ == '__main__':
    main()