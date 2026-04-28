"""
GadiW Migration — Step 2: Excel-driven Uploader
================================================
Reads the filled Excel from scan_usb.py and uploads files + metadata to Supabase.

Phase C1 (done): Login, validate, generate ambiguity report for client matching.
Phase C2 (now):  Actual file upload loop with retry + smart resume + orphan cleanup.

Usage:
    python migrate_to_gadiw.py --sheet <inventory.xlsx>
    python migrate_to_gadiw.py --sheet <inventory.xlsx> --resume
    python migrate_to_gadiw.py --sheet <inventory.xlsx> --dry-run

Options:
    --sheet <path>   Path to the filled Excel file (required)
    --resume         Skip rows already marked status='uploaded' in Excel
    --dry-run        Validate everything but don't upload (good for final check)
"""

import os
import sys
import argparse
import getpass
import json
import time
import secrets
from pathlib import Path
from datetime import datetime

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeRemainingColumn

console = Console()
load_dotenv()

# ============================================================================
# CONFIGURATION
# ============================================================================

SUPABASE_URL = os.getenv('VITE_SUPABASE_URL')
SUPABASE_ANON_KEY = os.getenv('VITE_SUPABASE_ANON_KEY')

if not SUPABASE_URL or not SUPABASE_ANON_KEY:
    console.print("[red]ERROR: Missing VITE_SUPABASE_URL or VITE_SUPABASE_ANON_KEY in .env[/red]")
    sys.exit(1)

REQUEST_TIMEOUT_SECONDS = 60
UPLOAD_TIMEOUT_SECONDS = 120
RETRY_ATTEMPTS = 3
RETRY_DELAY_SECONDS = 5
INTER_FILE_DELAY_MS = 50  # Throttle to stay under Supabase rate limits

DIRECTION_VALUES = {'פנימי', 'התקבל', 'נשלח'}
STORAGE_BUCKET = 'gadi-documents'

# Excel column indices (1-based)
COL_ROW_ID = 1
COL_FILE_PATH = 2
COL_FILE_NAME = 3
COL_FILE_SIZE_KB = 4
COL_FOLDER_HINT = 5
COL_FILE_MOD_DATE = 6
COL_CLIENT_NAME = 7
COL_DOC_TAG = 8
COL_DIRECTION = 9
COL_DOC_DATE = 10
COL_DESCRIPTION = 11
COL_STATUS = 12

# Status values written to Excel column L
STATUS_UPLOADED = 'uploaded'
STATUS_FAILED = 'failed'
STATUS_SKIPPED = 'skipped'


# ============================================================================
# AUTH
# ============================================================================

def login() -> dict:
    console.print()
    console.rule("[bold cyan]Login to Supabase[/bold cyan]")

    email = Prompt.ask("Email", default="avshi2@gmail.com")
    password = getpass.getpass("Password: ")

    url = f"{SUPABASE_URL}/auth/v1/token?grant_type=password"
    headers = {
        'apikey': SUPABASE_ANON_KEY,
        'Content-Type': 'application/json',
    }
    body = {'email': email, 'password': password}

    try:
        resp = requests.post(url, headers=headers, json=body, timeout=REQUEST_TIMEOUT_SECONDS)
    except requests.RequestException as e:
        console.print(f"[red]Network error during login: {e}[/red]")
        sys.exit(1)

    if resp.status_code != 200:
        console.print(f"[red]Login failed (HTTP {resp.status_code}): {resp.text[:300]}[/red]")
        sys.exit(1)

    data = resp.json()
    user_id = data.get('user', {}).get('id')
    access_token = data.get('access_token')

    if not user_id or not access_token:
        console.print(f"[red]Login response malformed: {data}[/red]")
        sys.exit(1)

    console.print(f"[green]✓ Logged in as {email}[/green]")
    console.print(f"  user_id: {user_id}")
    console.print()

    return {'email': email, 'user_id': user_id, 'access_token': access_token}


# ============================================================================
# FETCH CLIENTS
# ============================================================================

def fetch_all_clients(session: dict) -> list[dict]:
    url = f"{SUPABASE_URL}/rest/v1/clients?select=id,full_name&order=full_name.asc"
    headers = {
        'apikey': SUPABASE_ANON_KEY,
        'Authorization': f"Bearer {session['access_token']}",
        'Content-Type': 'application/json',
    }

    try:
        resp = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT_SECONDS)
    except requests.RequestException as e:
        console.print(f"[red]Network error fetching clients: {e}[/red]")
        sys.exit(1)

    if resp.status_code != 200:
        console.print(f"[red]Failed to fetch clients (HTTP {resp.status_code}): {resp.text[:300]}[/red]")
        sys.exit(1)

    clients = resp.json()
    console.print(f"[green]✓ Fetched {len(clients)} clients from Supabase[/green]")

    if clients:
        sample = ', '.join(c['full_name'] for c in clients[:5])
        more = f" (+{len(clients)-5} more)" if len(clients) > 5 else ""
        console.print(f"  Sample: {sample}{more}")
    console.print()

    return clients


# ============================================================================
# EXCEL READING + VALIDATION
# ============================================================================

class RowError:
    def __init__(self, row_num: int, file_name: str, problem: str, severity: str):
        self.row_num = row_num
        self.file_name = file_name
        self.problem = problem
        self.severity = severity


def load_excel_rows(sheet_path: Path) -> list[dict]:
    if not sheet_path.exists():
        console.print(f"[red]ERROR: Excel file not found: {sheet_path}[/red]")
        sys.exit(1)

    try:
        wb = load_workbook(sheet_path, data_only=True)
    except Exception as e:
        console.print(f"[red]Failed to open Excel: {e}[/red]")
        sys.exit(1)

    if 'Files' not in wb.sheetnames:
        console.print(f"[red]ERROR: No 'Files' sheet found in Excel.[/red]")
        sys.exit(1)

    ws = wb['Files']
    rows = []

    for row_num in range(2, ws.max_row + 1):
        if all(ws.cell(row=row_num, column=col).value is None for col in range(1, 13)):
            continue

        rows.append({
            'excel_row': row_num,
            'row_id': ws.cell(row=row_num, column=COL_ROW_ID).value,
            'file_path': ws.cell(row=row_num, column=COL_FILE_PATH).value,
            'file_name': ws.cell(row=row_num, column=COL_FILE_NAME).value,
            'file_size_kb': ws.cell(row=row_num, column=COL_FILE_SIZE_KB).value,
            'folder_hint': ws.cell(row=row_num, column=COL_FOLDER_HINT).value,
            'file_modified_date': ws.cell(row=row_num, column=COL_FILE_MOD_DATE).value,
            'client_name': ws.cell(row=row_num, column=COL_CLIENT_NAME).value,
            'doc_tag': ws.cell(row=row_num, column=COL_DOC_TAG).value,
            'direction': ws.cell(row=row_num, column=COL_DIRECTION).value,
            'doc_date': ws.cell(row=row_num, column=COL_DOC_DATE).value,
            'description': ws.cell(row=row_num, column=COL_DESCRIPTION).value,
            'status': ws.cell(row=row_num, column=COL_STATUS).value,
        })

    console.print(f"[green]✓ Loaded {len(rows)} data rows from Excel[/green]")
    console.print()
    return rows


def validate_and_match(rows: list[dict], clients: list[dict]) -> tuple[list[dict], list[RowError]]:
    valid_rows = []
    errors = []

    name_to_clients = {}
    for c in clients:
        key = (c['full_name'] or '').strip()
        name_to_clients.setdefault(key, []).append(c)

    for row in rows:
        row_num = row['excel_row']
        file_name = row['file_name'] or '(no name)'

        if not row['file_path']:
            errors.append(RowError(row_num, file_name, "missing file_path (open inventory.xlsx, column B should auto-fill)", "invalid"))
            continue

        client_name = (row['client_name'] or '').strip() if row['client_name'] else ''
        if not client_name:
            errors.append(RowError(row_num, file_name, "missing client_name (open inventory.xlsx, fill column G of this row)", "unmatched_client"))
            continue

        doc_tag = (row['doc_tag'] or '').strip() if row['doc_tag'] else ''
        if not doc_tag:
            errors.append(RowError(row_num, file_name, "missing doc_tag (open inventory.xlsx, fill column H of this row)", "invalid"))
            continue

        direction = (row['direction'] or '').strip() if row['direction'] else ''
        if direction not in DIRECTION_VALUES:
            errors.append(RowError(
                row_num, file_name,
                f"direction='{direction}' invalid (open inventory.xlsx, fix column I — must be: פנימי / התקבל / נשלח)",
                "invalid"
            ))
            continue

        matches = name_to_clients.get(client_name, [])
        if len(matches) == 0:
            suggestions = [
                c['full_name'] for c in clients
                if client_name in (c['full_name'] or '') or (c['full_name'] or '') in client_name
            ][:3]
            sugg_str = f" — did you mean: {', '.join(suggestions)}?" if suggestions else ""
            errors.append(RowError(
                row_num, file_name,
                f"client_name='{client_name}' not found in DB{sugg_str}",
                "unmatched_client"
            ))
            continue
        if len(matches) > 1:
            errors.append(RowError(
                row_num, file_name,
                f"client_name='{client_name}' matches {len(matches)} clients (ambiguous)",
                "ambiguous_client"
            ))
            continue

        row['client_id'] = matches[0]['id']
        valid_rows.append(row)

    return valid_rows, errors


# ============================================================================
# CLIENT REVIEW EXCEL
# ============================================================================

def generate_client_review(errors: list[RowError], clients: list[dict], output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Needs Review"
    ws.sheet_view.rightToLeft = True

    headers = ['excel_row', 'file_name', 'severity', 'problem', 'suggested_fix']
    header_font = Font(bold=True, color='FFFFFF', size=11)

    severity_colors = {
        'invalid': 'B71C1C',
        'unmatched_client': 'E65100',
        'ambiguous_client': 'FFA000',
    }

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor='1F3A5F')
        cell.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A2'

    for idx, err in enumerate(errors, 2):
        ws.cell(row=idx, column=1, value=err.row_num)
        ws.cell(row=idx, column=2, value=err.file_name)
        sev_cell = ws.cell(row=idx, column=3, value=err.severity)
        sev_cell.fill = PatternFill('solid', fgColor=severity_colors.get(err.severity, 'CCCCCC'))
        sev_cell.font = Font(color='FFFFFF', bold=True)
        sev_cell.alignment = Alignment(horizontal='center')
        ws.cell(row=idx, column=4, value=err.problem)
        if err.severity == 'unmatched_client':
            ws.cell(row=idx, column=5, value="Open inventory.xlsx, fix client_name in column G of this row")
        elif err.severity == 'ambiguous_client':
            ws.cell(row=idx, column=5, value="Use a more specific client_name (full unique name)")
        elif err.severity == 'invalid':
            ws.cell(row=idx, column=5, value="Fix the field mentioned in 'problem' column")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60

    ws_clients = wb.create_sheet("Available Clients in DB")
    ws_clients.sheet_view.rightToLeft = True
    ws_clients.cell(row=1, column=1, value='client_id').font = Font(bold=True, color='FFFFFF')
    ws_clients.cell(row=1, column=1).fill = PatternFill('solid', fgColor='1F3A5F')
    ws_clients.cell(row=1, column=2, value='full_name').font = Font(bold=True, color='FFFFFF')
    ws_clients.cell(row=1, column=2).fill = PatternFill('solid', fgColor='1F3A5F')

    for idx, c in enumerate(clients, 2):
        ws_clients.cell(row=idx, column=1, value=c['id'])
        ws_clients.cell(row=idx, column=2, value=c['full_name'])

    ws_clients.column_dimensions['A'].width = 40
    ws_clients.column_dimensions['B'].width = 40
    ws_clients.freeze_panes = 'A2'

    wb.save(output_path)


# ============================================================================
# UPLOAD LOOP — Phase C2 core
# ============================================================================

def make_safe_storage_path(file_name: str, user_id: str) -> str:
    """Returns ASCII-safe storage path, preserving original filename in metadata only."""
    timestamp = int(time.time() * 1000)
    random_suffix = secrets.token_hex(4)
    ext = Path(file_name).suffix.lower()
    # Strip non-ASCII from extension just in case
    ext = ''.join(c for c in ext if c.isascii() and (c.isalnum() or c == '.'))
    return f"inbox/{user_id}/{timestamp}_{random_suffix}{ext}"


def upload_one_row(row: dict, session: dict, log_lines: list, orphan_queue: list) -> tuple[bool, str]:
    """
    Returns (success, message).
    Implements 2-step transaction with retry + orphan cleanup.
    """
    file_path = Path(row['file_path'])
    user_id = session['user_id']
    access_token = session['access_token']

    # Verify file exists on disk
    if not file_path.exists():
        return False, f"file not found on disk: {file_path}"
    if not file_path.is_file():
        return False, f"path is not a file: {file_path}"

    storage_path = make_safe_storage_path(row['file_name'], user_id)
    storage_url = f"{SUPABASE_URL}/storage/v1/object/{STORAGE_BUCKET}/{storage_path}"

    # ===== Step 1: Upload to Storage with retry =====
    storage_succeeded = False
    last_error = ''

    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            with open(file_path, 'rb') as f:
                file_bytes = f.read()
        except OSError as e:
            return False, f"could not read file from disk: {e}"

        try:
            resp = requests.post(
                storage_url,
                headers={
                    'apikey': SUPABASE_ANON_KEY,
                    'Authorization': f'Bearer {access_token}',
                    'Content-Type': 'application/octet-stream',
                    'x-upsert': 'false',
                },
                data=file_bytes,
                timeout=UPLOAD_TIMEOUT_SECONDS,
            )
        except requests.RequestException as e:
            last_error = f"network error (attempt {attempt}/{RETRY_ATTEMPTS}): {e}"
            log_lines.append(f"  [retry] {last_error}")
            if attempt < RETRY_ATTEMPTS:
                time.sleep(RETRY_DELAY_SECONDS)
            continue

        if resp.status_code in (200, 201):
            storage_succeeded = True
            break
        else:
            last_error = f"storage HTTP {resp.status_code} (attempt {attempt}/{RETRY_ATTEMPTS}): {resp.text[:200]}"
            log_lines.append(f"  [retry] {last_error}")
            # 401/403 = auth, won't recover on retry
            if resp.status_code in (401, 403):
                return False, f"auth failure: {last_error}"
            if attempt < RETRY_ATTEMPTS:
                time.sleep(RETRY_DELAY_SECONDS)

    if not storage_succeeded:
        return False, f"storage upload failed after {RETRY_ATTEMPTS} attempts: {last_error}"

    # ===== Step 2: Insert documents row =====
    doc_date = row['doc_date']
    if doc_date is None or doc_date == '':
        # Fall back to file modified date
        doc_date = row['file_modified_date']

    # Convert to YYYY-MM-DD string if it's a datetime
    if hasattr(doc_date, 'strftime'):
        doc_date = doc_date.strftime('%Y-%m-%d')
    elif doc_date:
        doc_date = str(doc_date).strip()
    else:
        doc_date = None

    description = row['description']
    if description:
        description = str(description).strip() or None

    file_size_bytes = file_path.stat().st_size

    # Best-effort mime type from extension
    ext = file_path.suffix.lower()
    mime_map = {
        '.pdf': 'application/pdf',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.doc': 'application/msword',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.gif': 'image/gif',
        '.webp': 'image/webp', '.bmp': 'image/bmp',
        '.tiff': 'image/tiff', '.tif': 'image/tiff',
        '.txt': 'text/plain', '.rtf': 'application/rtf',
        '.eml': 'message/rfc822', '.msg': 'application/vnd.ms-outlook',
    }
    mime_type = mime_map.get(ext, 'application/octet-stream')

    doc_row = {
        'file_name': row['file_name'],
        'file_url': storage_path,
        'file_size': file_size_bytes,
        'mime_type': mime_type,
        'doc_tag': str(row['doc_tag']).strip(),
        'direction': str(row['direction']).strip(),
        'doc_date': doc_date,
        'description': description,
        'client_id': row['client_id'],
    }

    insert_url = f"{SUPABASE_URL}/rest/v1/documents"

    insert_succeeded = False
    last_error = ''

    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            resp = requests.post(
                insert_url,
                headers={
                    'apikey': SUPABASE_ANON_KEY,
                    'Authorization': f'Bearer {access_token}',
                    'Content-Type': 'application/json',
                    'Prefer': 'return=minimal',
                },
                json=doc_row,
                timeout=REQUEST_TIMEOUT_SECONDS,
            )
        except requests.RequestException as e:
            last_error = f"DB network error (attempt {attempt}/{RETRY_ATTEMPTS}): {e}"
            log_lines.append(f"  [retry] {last_error}")
            if attempt < RETRY_ATTEMPTS:
                time.sleep(RETRY_DELAY_SECONDS)
            continue

        if resp.status_code in (200, 201, 204):
            insert_succeeded = True
            break
        else:
            last_error = f"DB HTTP {resp.status_code} (attempt {attempt}/{RETRY_ATTEMPTS}): {resp.text[:200]}"
            log_lines.append(f"  [retry] {last_error}")
            if resp.status_code in (401, 403):
                # Schedule orphan cleanup, don't keep retrying auth failure
                orphan_queue.append({'storage_path': storage_path, 'reason': 'auth failure on DB insert'})
                return False, f"auth failure: {last_error}"
            if attempt < RETRY_ATTEMPTS:
                time.sleep(RETRY_DELAY_SECONDS)

    if not insert_succeeded:
        # Schedule orphan cleanup — Storage POST succeeded but DB INSERT failed
        orphan_queue.append({'storage_path': storage_path, 'reason': last_error})
        return False, f"DB insert failed (orphan in storage): {last_error}"

    return True, "ok"


def cleanup_orphans(orphan_queue: list, session: dict, log_lines: list) -> tuple[int, int]:
    """Returns (cleaned, still_orphan)."""
    if not orphan_queue:
        return 0, 0

    console.print()
    console.print(f"[yellow]Cleaning up {len(orphan_queue)} orphan storage objects...[/yellow]")

    cleaned = 0
    still_orphan = []

    for orphan in orphan_queue:
        storage_path = orphan['storage_path']
        delete_url = f"{SUPABASE_URL}/storage/v1/object/{STORAGE_BUCKET}/{storage_path}"

        try:
            resp = requests.delete(
                delete_url,
                headers={
                    'apikey': SUPABASE_ANON_KEY,
                    'Authorization': f"Bearer {session['access_token']}",
                },
                timeout=REQUEST_TIMEOUT_SECONDS,
            )
            if resp.status_code in (200, 204):
                cleaned += 1
                log_lines.append(f"  [cleanup] deleted orphan: {storage_path}")
            else:
                still_orphan.append(orphan)
                log_lines.append(f"  [cleanup-FAILED] {storage_path}: HTTP {resp.status_code}")
        except requests.RequestException as e:
            still_orphan.append(orphan)
            log_lines.append(f"  [cleanup-FAILED] {storage_path}: {e}")

    if still_orphan:
        console.print(f"[red]⚠ {len(still_orphan)} orphans could NOT be cleaned. See log file.[/red]")
        console.print(f"[red]  Run reconciliation SQL after migration to detect them.[/red]")

    return cleaned, len(still_orphan)


def write_status_to_excel(sheet_path: Path, excel_row: int, status: str, message: str):
    """Updates column L (status) for one row, opens+saves Excel each time."""
    try:
        wb = load_workbook(sheet_path)
        ws = wb['Files']
        # Truncate message to fit cell
        full_status = status if not message or status == STATUS_UPLOADED else f"{status}: {message[:100]}"
        ws.cell(row=excel_row, column=COL_STATUS, value=full_status)

        # Color the status cell
        if status == STATUS_UPLOADED:
            ws.cell(row=excel_row, column=COL_STATUS).fill = PatternFill('solid', fgColor='C8E6C9')
        elif status == STATUS_FAILED:
            ws.cell(row=excel_row, column=COL_STATUS).fill = PatternFill('solid', fgColor='FFCDD2')
        elif status == STATUS_SKIPPED:
            ws.cell(row=excel_row, column=COL_STATUS).fill = PatternFill('solid', fgColor='E0E0E0')

        wb.save(sheet_path)
    except Exception as e:
        # Don't crash migration over status writeback
        console.print(f"[yellow]Warning: could not write status to Excel row {excel_row}: {e}[/yellow]")


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description='GadiW migration uploader')
    parser.add_argument('--sheet', required=True, help='Path to filled inventory Excel')
    parser.add_argument('--resume', action='store_true', help='Skip rows with status=uploaded')
    parser.add_argument('--dry-run', action='store_true', help='Validate only, no uploads')
    args = parser.parse_args()

    sheet_path = Path(args.sheet).resolve()

    console.rule("[bold cyan]GadiW Migration Uploader[/bold cyan]")
    console.print(f"Excel: [bold]{sheet_path}[/bold]")
    if args.resume:
        console.print(f"[yellow]RESUME MODE: skipping rows with status='uploaded'[/yellow]")
    if args.dry_run:
        console.print(f"[yellow]DRY RUN: no actual uploads will happen[/yellow]")
    console.print()

    # Validate file exists BEFORE asking for password
    if not sheet_path.exists():
        console.print(f"[red]ERROR: File not found: {sheet_path}[/red]")
        console.print(f"[yellow]Available Excel files in this folder:[/yellow]")
        for f in Path.cwd().glob('*.xlsx'):
            console.print(f"  • {f.name}")
        sys.exit(1)

    if not sheet_path.is_file():
        console.print(f"[red]ERROR: Path is not a file: {sheet_path}[/red]")
        sys.exit(1)

    # Step 1: Login
    session = login()

    # Step 2: Fetch clients
    clients = fetch_all_clients(session)
    if not clients:
        console.print("[red]ERROR: No clients found in DB. Create at least 1 client first.[/red]")
        sys.exit(1)

    # Step 3: Load Excel
    rows = load_excel_rows(sheet_path)
    if not rows:
        console.print("[red]ERROR: No data rows in Excel.[/red]")
        sys.exit(1)

    # Step 4: Validate + match
    console.print("[cyan]Validating rows...[/cyan]")
    valid_rows, errors = validate_and_match(rows, clients)

    # Summary table
    table = Table(title="Validation Summary", show_header=True, header_style="bold cyan")
    table.add_column("Metric", style="cyan")
    table.add_column("Count", justify="right")
    table.add_row("Total rows in Excel", str(len(rows)))
    table.add_row("[green]Valid (ready to upload)[/green]", f"[green]{len(valid_rows)}[/green]")
    table.add_row("[yellow]Need review[/yellow]", f"[yellow]{len(errors)}[/yellow]")
    console.print(table)
    console.print()

    if errors:
        sev_counts = {}
        for e in errors:
            sev_counts[e.severity] = sev_counts.get(e.severity, 0) + 1
        sev_table = Table(title="Error Breakdown", show_header=True, header_style="bold yellow")
        sev_table.add_column("Severity", style="yellow")
        sev_table.add_column("Count", justify="right")
        for sev, count in sorted(sev_counts.items(), key=lambda x: -x[1]):
            sev_table.add_row(sev, str(count))
        console.print(sev_table)
        console.print()

        timestamp = datetime.now().strftime('%d%m%Y_%H%M%S')
        review_path = sheet_path.parent / f'client_review_{timestamp}.xlsx'
        generate_client_review(errors, clients, review_path)

        console.print(f"[yellow]⚠ {len(errors)} rows need review. Cannot proceed with upload.[/yellow]")
        console.print(f"[yellow]  → Review file: {review_path.name}[/yellow]")
        console.print(f"[yellow]  → Fix the rows in {sheet_path.name}[/yellow]")
        console.print(f"[yellow]  → Re-run this command.[/yellow]")
        console.print()
        sys.exit(0)

    console.print(f"[bold green]✓ All {len(valid_rows)} rows valid.[/bold green]")
    console.print()

    if args.dry_run:
        console.print("[yellow]DRY RUN — would upload these rows but stopping here.[/yellow]")
        sys.exit(0)

    # Step 5: Resume filtering
    if args.resume:
        before = len(valid_rows)
        valid_rows = [r for r in valid_rows if r.get('status') != STATUS_UPLOADED]
        skipped = before - len(valid_rows)
        if skipped:
            console.print(f"[cyan]RESUME: skipping {skipped} already-uploaded rows.[/cyan]")
        console.print(f"[cyan]Will upload: {len(valid_rows)} rows.[/cyan]")
        console.print()

    if not valid_rows:
        console.print("[green]Nothing to upload. All rows already done. ✓[/green]")
        sys.exit(0)

    # Step 6: Confirm before starting
    console.print(f"[bold]About to upload {len(valid_rows)} files to Supabase.[/bold]")
    response = Prompt.ask("Proceed?", choices=['y', 'n'], default='y')
    if response != 'y':
        console.print("[red]Aborted by user.[/red]")
        sys.exit(0)
    console.print()

    # Step 7: Upload loop
    log_lines = [f"Migration log — started {datetime.now().isoformat()}"]
    orphan_queue = []
    success_count = 0
    failure_count = 0

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TimeRemainingColumn(),
        console=console,
    ) as progress:
        task = progress.add_task(f"Uploading {len(valid_rows)} files...", total=len(valid_rows))

        for idx, row in enumerate(valid_rows, 1):
            file_name = row['file_name'] or '(no name)'
            progress.update(task, description=f"[{idx}/{len(valid_rows)}] {file_name[:50]}")

            log_lines.append(f"\n[{idx}/{len(valid_rows)}] {file_name} (excel row {row['excel_row']})")
            log_lines.append(f"  client_id: {row['client_id']}")
            log_lines.append(f"  doc_tag: {row['doc_tag']}")

            success, message = upload_one_row(row, session, log_lines, orphan_queue)

            if success:
                success_count += 1
                log_lines.append(f"  [OK] uploaded")
                write_status_to_excel(sheet_path, row['excel_row'], STATUS_UPLOADED, '')
            else:
                failure_count += 1
                log_lines.append(f"  [FAIL] {message}")
                write_status_to_excel(sheet_path, row['excel_row'], STATUS_FAILED, message)

            progress.advance(task)

            # Rate-limit throttle
            time.sleep(INTER_FILE_DELAY_MS / 1000.0)

    # Step 8: Cleanup orphans
    cleaned, still_orphan = cleanup_orphans(orphan_queue, session, log_lines)

    # Step 9: Save log
    timestamp = datetime.now().strftime('%d%m%Y_%H%M%S')
    log_path = sheet_path.parent / f'migration_log_{timestamp}.txt'
    log_path.write_text('\n'.join(log_lines), encoding='utf-8')

    # Step 10: Final summary
    console.print()
    console.rule("[bold green]Migration Complete[/bold green]")

    summary = Table(show_header=True, header_style="bold cyan")
    summary.add_column("Result", style="cyan")
    summary.add_column("Count", justify="right")
    summary.add_row("[green]✓ Uploaded successfully[/green]", f"[green]{success_count}[/green]")
    summary.add_row("[red]✗ Failed[/red]", f"[red]{failure_count}[/red]")
    summary.add_row("[yellow]Orphans cleaned[/yellow]", f"[yellow]{cleaned}[/yellow]")
    if still_orphan:
        summary.add_row("[red]Orphans REMAINING (need manual cleanup)[/red]", f"[red]{still_orphan}[/red]")
    console.print(summary)
    console.print()
    console.print(f"[bold]Log file:[/bold] {log_path.name}")
    console.print(f"[bold]Excel updated[/bold] with status in column L")
    console.print()

    if failure_count > 0:
        console.print(f"[yellow]To retry failed rows:[/yellow]")
        console.print(f"[yellow]  python migrate_to_gadiw.py --sheet {sheet_path.name} --resume[/yellow]")
        console.print()

    if still_orphan > 0:
        console.print(f"[red]⚠ {still_orphan} orphan files still in Supabase Storage.[/red]")
        console.print(f"[red]  Run reconciliation SQL after migration to detect + clean.[/red]")
        console.print()


if __name__ == '__main__':
    main()